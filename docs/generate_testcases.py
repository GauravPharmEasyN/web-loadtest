#!/usr/bin/env python3
"""
Generate Doc Consult test cases Excel from PRD and screenshots.

Creates docs/DocConsult_TestCases.xlsx with sheets:
- Summary
- Phase1_Discovery
- Phase1_PostOrder
- Stage2_ConsultFlow
- Phase3_Prescriptions
- NonProduct_Analytics

Columns:
ID, Title, Objective, Preconditions, Steps, Expected Result, Priority,
Type, Module, Screen, AB Variant, Data, Image Ref, Owner, Tags
"""

from pathlib import Path
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


ROOT = Path(__file__).resolve().parents[1]
OUT_XLSX = ROOT / "docs" / "DocConsult_TestCases.xlsx"


Row = Dict[str, str]


def _header() -> List[str]:
    return [
        "ID",
        "Title",
        "Objective",
        "Preconditions",
        "Steps",
        "Expected Result",
        "Priority",
        "Type",
        "Module",
        "Screen",
        "AB Variant",
        "Data",
        "Image Ref",
        "Owner",
        "Tags",
    ]


def add_sheet(wb: Workbook, title: str, rows: List[Row]) -> None:
    ws = wb.create_sheet(title=title)
    headers = _header()
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    # Auto width (simple heuristic)
    for i, h in enumerate(headers, start=1):
        max_len = max(len(str(ws.cell(row=r, column=i).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 60)


def phase1_discovery_rows() -> List[Row]:
    rows: List[Row] = []

    def r(idx: int, title: str, objective: str, steps: str, expected: str, screen: str, ab: str = "A/B Eligible", img: str = "") -> Row:
        return {
            "ID": f"P1-DISC-{idx:03d}",
            "Title": title,
            "Objective": objective,
            "Preconditions": "User is logged in; app build with feature flags active; marketing panel config available",
            "Steps": steps,
            "Expected Result": expected,
            "Priority": "P1",
            "Type": "Functional",
            "Module": "Diagnostics",
            "Screen": screen,
            "AB Variant": ab,
            "Data": "N/A",
            "Image Ref": img,
            "Owner": "QA",
            "Tags": "discovery,doc-consult,ab-test",
        }

    rows.append(
        r(
            1,
            "Homepage sliding popup shows Doc Consult",
            "Validate sliding popup visibility and content from marketing config",
            "Launch app > Navigate to Diagnostics Home > Observe top sliding popup",
            "Popup is visible; contains image, headline, and copy; dismiss CTA works; impression event fired",
            "Diagnostics Home",
            img="s1_home_popup.png",
        )
    )
    rows.append(
        r(
            2,
            "Popup controllable via marketing panel",
            "Verify content changes propagate without app update",
            "Toggle headline/image via marketing panel > Refresh Home",
            "New content reflected within TTL; no layout break",
            "Diagnostics Home",
        )
    )
    rows.append(
        r(
            3,
            "A/B flag routes user to variant",
            "Confirm equal bucketing and sticky assignment",
            "Enable experiment; log user bucket; relaunch app",
            "User remains in same bucket; events capture variant",
            "Diagnostics Home",
            ab="Variant A/B",
        )
    )
    rows.append(
        r(
            4,
            "Search page badge for Doc Consult",
            "Ensure SRP shows consult advantage row",
            "Search for 'Thyroid' > Observe SRP banners",
            "Consult advantage tile/badge visible; CTR captured",
            "Search Results",
            img="s2_search_badge.png",
        )
    )
    rows.append(
        r(
            5,
            "PDP consult strip present (non-ECG)",
            "Show single standardized strip across packages except ECG",
            "Open non-ECG PDP > scroll to consult strip",
            "Strip visible; says consult free post reports; event logged",
            "PDP",
        )
    )
    rows.append(
        r(
            6,
            "ECG PDP hides consult strip",
            "Do not show consult on ECG",
            "Open ECG PDP",
            "No consult strip present",
            "PDP",
            ab="N/A",
        )
    )
    rows.append(
        r(
            7,
            "Cart page consult SKU appears",
            "Validate consult as SKU with remove option",
            "Add lab package to cart > open cart",
            "Consult SKU shows with MRP/discount/free; removable; priced by cohort",
            "Cart",
            img="s3_cart_consult_sku.png",
        )
    )
    rows.append(
        r(
            8,
            "Cart consult pricing rules",
            "Ensure cohort/avg order price controls apply",
            "Load user in free cohort > open cart",
            "Consult SKU shows ₹0 or discounted as per rules; fallback to ₹0 on price failure",
            "Cart",
        )
    )
    return rows


def phase1_postorder_rows() -> List[Row]:
    rows: List[Row] = []

    def r(idx: int, title: str, objective: str, steps: str, expected: str, img: str = "") -> Row:
        return {
            "ID": f"P1-POST-{idx:03d}",
            "Title": title,
            "Objective": objective,
            "Preconditions": "Order placed; at least one patient; app has post-order screens",
            "Steps": steps,
            "Expected Result": expected,
            "Priority": "P1",
            "Type": "Functional",
            "Module": "Post-Order",
            "Screen": "Order Details",
            "AB Variant": "N/A",
            "Data": "N/A",
            "Image Ref": img,
            "Owner": "QA",
            "Tags": "post-order,coupon",
        }

    rows.append(
        r(
            1,
            "Consult banner below lab reports",
            "Show banner before report generation and after",
            "Open order details before reports ready; then after ready",
            "Banner visible both contexts with correct messaging",
            img="s4_order_banner.png",
        )
    )
    rows.append(
        r(
            2,
            "Coupon generation when signed report available",
            "Verify coupon created only when digitally signed report exists",
            "Mark one patient report as PE-signed > refresh",
            "Coupon generated; validity 1 month; unique per user",
        )
    )
    rows.append(
        r(
            3,
            "Coupon expiry handling",
            "Backend-driven expiry respected",
            "Set coupon expired > open order details",
            "Banner shows expired; CTA disabled or shows new ETA",
        )
    )
    rows.append(
        r(
            4,
            "Copy coupon interaction",
            "Copy-to-clipboard feedback",
            "Tap Copy on coupon card",
            "Toast: 'Coupon copied'; value visible",
        )
    )
    rows.append(
        r(
            5,
            "Book consult deep link",
            "Deep link contains session tokens and navigates",
            "Tap 'Consult a doctor now'",
            "Navigated to Doc Consult selection with user validated",
        )
    )
    return rows


def stage2_consult_rows() -> List[Row]:
    rows: List[Row] = []

    def r(idx: int, title: str, steps: str, expected: str, data: str = "") -> Row:
        return {
            "ID": f"S2-CONS-{idx:03d}",
            "Title": title,
            "Objective": "Validate free consult selection and order creation",
            "Preconditions": "Deep link from post-order; same user session; coupon or wallet credit available",
            "Steps": steps,
            "Expected Result": expected,
            "Priority": "P1",
            "Type": "Functional",
            "Module": "Doc Consult",
            "Screen": "Consult Listing/Cart",
            "AB Variant": "N/A",
            "Data": data,
            "Image Ref": "s5_consult_checkout.png",
            "Owner": "QA",
            "Tags": "consult,free,coupon,security",
        }

    rows.append(
        r(
            1,
            "Deep link validates same user",
            "Open deep link with same-account token",
            "Consult opens; user auto-authenticated; no cross-user access",
        )
    )
    rows.append(
        r(
            2,
            "Token misuse prevention",
            "Share deep link to another user/device and open",
            "Access denied; prompts login as original user; no free consult applied",
        )
    )
    rows.append(
        r(
            3,
            "Auto-apply coupon shows price ₹0",
            "Navigate to consult checkout",
            "Coupon auto-applied; payable amount ₹0; GST 0; CTA enabled",
        )
    )
    rows.append(
        r(
            4,
            "Fallback price to ₹0 on calc failure",
            "Mock pricing service failure",
            "Price displays ₹0.0; order continues",
        )
    )
    rows.append(
        r(
            5,
            "Wallet-credit approach supported",
            "Enable hidden wallet credit; proceed to checkout",
            "Visible price reduced to 0 via credits; order placed",
        )
    )
    rows.append(
        r(
            6,
            "Consult modes selectable",
            "Choose Audio then Video and book",
            "Selection persists; order success",
        )
    )
    return rows


def phase3_prescriptions_rows() -> List[Row]:
    rows: List[Row] = []

    def r(idx: int, title: str, steps: str, expected: str, tags: str = "prescriptions,conversion") -> Row:
        return {
            "ID": f"P3-PRES-{idx:03d}",
            "Title": title,
            "Objective": "Validate post-prescription lab test identification and conversion",
            "Preconditions": "User uploads prescription or doctor generates post-consult",
            "Steps": steps,
            "Expected Result": expected,
            "Priority": "P2",
            "Type": "Functional",
            "Module": "Diagnostics + Consult",
            "Screen": "Prescription/Recommendations",
            "AB Variant": "N/A",
            "Data": "N/A",
            "Image Ref": "",
            "Owner": "QA",
            "Tags": tags,
        }

    rows.append(
        r(
            1,
            "Identify lab tests from prescription",
            "Upload Rx > system parses and suggests tests",
            "Relevant tests surfaced; accuracy acceptable; events logged",
        )
    )
    rows.append(
        r(
            2,
            "Doctor recommends retest and adds to cart",
            "Doctor workflow adds tests; user sees cart",
            "Cart shows recommended tests and total; user can place order",
        )
    )
    return rows


def analytics_rows() -> List[Row]:
    rows: List[Row] = []

    def r(idx: int, title: str, steps: str, expected: str, tags: str) -> Row:
        return {
            "ID": f"ANL-{idx:03d}",
            "Title": title,
            "Objective": "Ensure analytics and guardrails",
            "Preconditions": "Analytics SDK configured; experiment flags set",
            "Steps": steps,
            "Expected Result": expected,
            "Priority": "P1",
            "Type": "Analytics",
            "Module": "All",
            "Screen": "Multiple",
            "AB Variant": "N/A",
            "Data": "N/A",
            "Image Ref": "",
            "Owner": "QA",
            "Tags": tags,
        }

    rows.append(
        r(
            1,
            "Key funnel metrics captured",
            "Trigger impressions, clicks, add-to-cart, orders across variants",
            "Events emitted with properties: variant, cohort, coupon_applied, price, user_id",
            "metrics,funnel",
        )
    )
    rows.append(
        r(
            2,
            "User safety copy tone",
            "Check UI copy for selling tone",
            "Copy emphasizes service/help; no aggressive upsell",
            "ux,safety",
        )
    )
    rows.append(
        r(
            3,
            "Spam/fraud prevention",
            "Attempt multiple free consults via replays",
            "Rate limits enforced; single redemption per order/user",
            "security,fraud",
        )
    )
    return rows


def build_workbook() -> Workbook:
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    # Summary
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Doc Consult Test Cases"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A3"] = "Generated from PRD and screenshots. Update Image Ref with actual paths if embedding."
    ws.column_dimensions["A"].width = 120

    add_sheet(wb, "Phase1_Discovery", phase1_discovery_rows())
    add_sheet(wb, "Phase1_PostOrder", phase1_postorder_rows())
    add_sheet(wb, "Stage2_ConsultFlow", stage2_consult_rows())
    add_sheet(wb, "Phase3_Prescriptions", phase3_prescriptions_rows())
    add_sheet(wb, "NonProduct_Analytics", analytics_rows())

    return wb


def main() -> None:
    wb = build_workbook()
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"Wrote: {OUT_XLSX}")


if __name__ == "__main__":
    main()



